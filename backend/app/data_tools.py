import csv
import io
import zipfile
from datetime import date, datetime
from fastapi import HTTPException
from sqlalchemy import text, Integer, Float, Boolean, Date, DateTime
from . import models

TABLE_MODELS = {
    "planos": models.Plano,
    "marcas": models.Marca,
    "estados": models.Estado,
    "tipos_host": models.TipoHost,
    "tipos_servidor": models.TipoServidor,
    "subestaciones": models.Subestacion,
    "blindobarras": models.Blindobarra,
    "hosts": models.Host,
    "racks": models.Rack,
    "aplicaciones": models.Aplicacion,
    "dependencias_app_servidor": models.DependenciaAppHost,
    "procesos_planta": models.ProcesoPlanta,
    "catalogo_equipos": models.CatalogoEquipo,
    "stock_consumibles": models.StockConsumible,
}

def parse_date(val):
    if not val:
        return None
    if isinstance(val, (date, datetime)):
        return val
    val = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None

def parse_int(val):
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None

def parse_float(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def parse_bool(val):
    if val is None or val == "":
        return None
    val_str = str(val).strip().lower()
    if val_str in ("true", "1", "t", "y", "yes", "sí", "si"):
        return True
    if val_str in ("false", "0", "f", "n", "no"):
        return False
    return None

def cast_value(col, val):
    col_type = col.type
    if val is None:
        return None
    
    if isinstance(col_type, Integer):
        return parse_int(val)
    elif isinstance(col_type, Float):
        return parse_float(val)
    elif isinstance(col_type, Boolean):
        return parse_bool(val)
    elif isinstance(col_type, Date):
        return parse_date(val)
    elif isinstance(col_type, DateTime):
        if isinstance(val, datetime):
            return val
        if isinstance(val, date):
            return datetime(val.year, val.month, val.day)
        val_str = str(val).strip()
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
        d = parse_date(val)
        if d:
            return datetime(d.year, d.month, d.day)
        return None
    else:
        return str(val)

def read_xlsx_to_dicts(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="La biblioteca 'openpyxl' no está instalada en el servidor backend."
        )
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    rows = list(sheet.rows)
    if not rows:
        return []
    
    headers = [cell.value for cell in rows[0]]
    # Ensure headers are strings
    headers = [str(h).strip() if h is not None else "" for h in headers]
    
    dicts = []
    for row in rows[1:]:
        row_values = [cell.value for cell in row]
        if not any(v is not None for v in row_values):
            continue
        row_dict = {}
        for h, v in zip(headers, row_values):
            if h:
                row_dict[h] = v
        dicts.append(row_dict)
    return dicts

def read_csv_to_dicts(file_bytes):
    text = file_bytes.decode('utf-8-sig', errors='ignore')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)

def upsert_row(db, model, row_dict):
    columns = model.__table__.columns
    
    # Cast row values according to database types
    cleaned_row = {}
    for col_name, col in columns.items():
        if col_name in row_dict:
            cleaned_row[col_name] = cast_value(col, row_dict[col_name])
            
    # Try to find existing record
    existing = None
    
    # 1. Match by primary key id if provided
    if "id" in cleaned_row and cleaned_row["id"] is not None:
        existing = db.query(model).filter(model.id == cleaned_row["id"]).first()
        
    # 2. Match by unique name if model has 'nombre' column
    if not existing and "nombre" in columns.keys() and cleaned_row.get("nombre"):
        existing = db.query(model).filter(model.nombre == cleaned_row["nombre"]).first()
        
    # 3. Match associations by key fields
    if not existing and model.__tablename__ == "stock_consumibles" and cleaned_row.get("catalogo_id") and cleaned_row.get("ubicacion"):
        existing = db.query(model).filter(
            model.catalogo_id == cleaned_row["catalogo_id"],
            model.ubicacion == cleaned_row["ubicacion"]
        ).first()
        
    if not existing and model.__tablename__ == "dependencias_app_servidor" and cleaned_row.get("app_id") and cleaned_row.get("host_id"):
        existing = db.query(model).filter(
            model.app_id == cleaned_row["app_id"],
            model.host_id == cleaned_row["host_id"]
        ).first()

    if not existing and model.__tablename__ == "procesos_planta" and cleaned_row.get("nombre_proceso") and cleaned_row.get("aplicacion_id"):
        existing = db.query(model).filter(
            model.nombre_proceso == cleaned_row["nombre_proceso"],
            model.aplicacion_id == cleaned_row["aplicacion_id"]
        ).first()

    if existing:
        for key, val in cleaned_row.items():
            if key != "id":
                setattr(existing, key, val)
        return "updated"
    else:
        new_record = model(**cleaned_row)
        db.add(new_record)
        return "inserted"

def reset_db_sequences(db):
    for table_name in TABLE_MODELS.keys():
        try:
            db.execute(text(f"""
                SELECT setval(
                    pg_get_serial_sequence('{table_name}', 'id'), 
                    COALESCE(MAX(id), 1)
                ) FROM {table_name};
            """))
            db.commit()
        except Exception:
            db.rollback()

def export_table_to_csv(db, model):
    output = io.StringIO()
    columns = model.__table__.columns.keys()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    
    records = db.query(model).order_by(model.id).all()
    for record in records:
        row = {}
        for col in columns:
            val = getattr(record, col)
            if isinstance(val, (date, datetime)):
                row[col] = val.isoformat()
            else:
                row[col] = val
        writer.writerow(row)
    return output.getvalue()

def export_table_to_xlsx(db, model):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="La biblioteca 'openpyxl' no está instalada en el servidor backend."
        )
    wb = Workbook()
    sheet = wb.active
    columns = list(model.__table__.columns.keys())
    
    sheet.append(columns)
    records = db.query(model).order_by(model.id).all()
    for record in records:
        row = []
        for col in columns:
            val = getattr(record, col)
            if isinstance(val, (date, datetime)):
                row.append(val.isoformat())
            else:
                row.append(val)
        sheet.append(row)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def export_all_tables_to_zip(db):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for table_name, model in TABLE_MODELS.items():
            csv_content = export_table_to_csv(db, model)
            zip_file.writestr(f"{table_name}.csv", csv_content)
    return zip_buffer.getvalue()

def get_sample_template(table_name):
    model = TABLE_MODELS.get(table_name)
    if not model:
        raise HTTPException(status_code=404, detail="Tabla no encontrada")
    
    output = io.StringIO()
    columns = model.__table__.columns.keys()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    
    # Generate some simple mockup/example rows for context
    mock_row = {}
    for col_name, col in model.__table__.columns.items():
        if col_name == "id":
            mock_row[col_name] = "1"
        elif col_name == "nombre":
            mock_row[col_name] = f"Ejemplo_{table_name}"
        elif isinstance(col.type, Integer):
            mock_row[col_name] = "10"
        elif isinstance(col.type, Float):
            mock_row[col_name] = "220.5"
        elif isinstance(col.type, Boolean):
            mock_row[col_name] = "true"
        elif isinstance(col.type, Date):
            mock_row[col_name] = "2026-08-12"
        elif isinstance(col.type, DateTime):
            mock_row[col_name] = "2026-08-12 12:00:00"
        else:
            mock_row[col_name] = "Texto Muestra"
            
    writer.writerow(mock_row)
    return output.getvalue()
